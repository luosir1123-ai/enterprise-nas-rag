"""Build a row-level SQLite sidecar index for selected RAGFlow Excel files."""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import sqlite3
import tempfile
import unicodedata
from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
from typing import Any

from business_metadata import extract_models, infer_business_metadata


SCHEMA_VERSION = "excel-row-index-v2"
DEFAULT_CONFIG_PATH = Path(os.getenv("RAGFLOW_EXCEL_ROW_INDEX_CONFIG", "/tmp/excel_row_index.json"))
DEFAULT_DATABASE_PATH = Path(
    os.getenv("RAGFLOW_EXCEL_ROW_INDEX_PATH", "/ragflow/structured-index/excel_rows.sqlite3")
)
DEFAULT_REPORT_PATH = Path(
    os.getenv("RAGFLOW_EXCEL_ROW_INDEX_REPORT_PATH", "/tmp/ragflow_excel_row_index_report.json")
)

HEADER_ALIASES = {
    "model": {"型号", "model", "modelno", "modelnumber", "货号", "itemno", "产品型号"},
    "serial_number": {"序号", "no", "number", "编号"},
    "product": {"品名", "product", "产品", "产品名称", "productname"},
    "product_image": {"产品图片", "图片", "image", "photo"},
    "feature": {"卖点", "feature", "features", "产品卖点"},
    "specification": {"描述", "specification", "specifications", "spec", "参数", "规格"},
    "moq": {"moq", "minimumorderquantity", "最小起订量", "起订量"},
    "remarks": {"备注", "remark", "remarks", "note", "notes"},
    "supplier": {"供应商", "supplier", "vendor", "工厂"},
    "cost_price": {"成本价", "成本", "cost", "costprice"},
    "final_price": {"终报价", "最终报价", "finalprice", "quotedprice"},
    "profit": {"profit", "利润", "毛利", "margin"},
    "currency": {"currency", "币种", "货币"},
    "tax_status": {"tax", "taxstatus", "税", "税率", "含税", "未税"},
    "quantity": {"quantity", "qty", "数量"},
    "country_standard": {"country", "region", "国家", "地区", "国家版本", "规格版本"},
}
MODEL_FIELD_KEYS = {"model"}
UNIT_RE = re.compile(r"^\s*([￥¥$]?)\s*(-?\d+(?:\.\d+)?)\s*([A-Za-z%元个只套台件箱PCSpcs]*)\s*$")
RANGE_RE = re.compile(
    r"^\s*(?P<prefix>US\$|HK\$|RMB|CNY|USD|EUR|GBP|￥|¥|\$)?\s*"
    r"(?P<minimum>-?\d+(?:\.\d+)?)\s*(?:-|~|至|到)\s*"
    r"(?P<maximum>-?\d+(?:\.\d+)?)\s*(?P<unit>[A-Za-z%元个只套台件箱]*)\s*$",
    re.IGNORECASE,
)
CURRENCY_ALIASES = {
    "￥": "CNY", "¥": "CNY", "RMB": "CNY", "CNY": "CNY",
    "$": "USD", "US$": "USD", "USD": "USD", "HK$": "HKD",
    "EUR": "EUR", "GBP": "GBP",
}
UNIT_ALIASES = {
    "PCS": "PCS", "PC": "PCS", "件": "PCS", "个": "PCS", "只": "PCS",
    "套": "SET", "台": "UNIT", "箱": "BOX", "%": "%",
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        if value.is_integer():
            return str(int(value))
    text = unicodedata.normalize("NFKC", str(value)).replace("\x00", "")
    return re.sub(r"[ \t]+", " ", text).strip()


def normalize_header(value: Any) -> str:
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", normalize_text(value).casefold())


def canonical_header(value: Any) -> str:
    normalized = normalize_header(value)
    if not normalized:
        return ""
    for field_key, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return field_key
    if "成本" in normalized:
        return "cost_price"
    if "终报价" in normalized or "最终报价" in normalized:
        return "final_price"
    if "单价" in normalized or "报价" in normalized or normalized in {"price", "unitprice"}:
        return "unit_price"
    return ""


def detect_header(values: list[str]) -> list[str] | None:
    field_keys = [canonical_header(value) for value in values]
    recognized = [field_key for field_key in field_keys if field_key]
    if len(recognized) < 2:
        return None
    if "model" not in recognized and len(recognized) < 4:
        return None
    return field_keys


def extract_row_models(value: str) -> list[str]:
    normalized = normalize_text(value)
    compact = re.sub(r"(?<=[_-])\s+", "", normalized)
    compact = re.sub(r"\s+(?=[_-])", "", compact)
    models = extract_models(compact)
    return [
        model
        for model in models
        if not any(other != model and other.endswith(model) for other in models)
    ]


def parse_numeric(value: str) -> tuple[float | None, str]:
    match = UNIT_RE.match(normalize_text(value))
    if not match:
        return None, ""
    prefix, number, suffix = match.groups()
    unit = f"{prefix}{suffix}".upper()
    try:
        return float(number), unit
    except ValueError:
        return None, unit


def parse_business_value(value: str, field_name: str = "", field_key: str = "") -> dict[str, Any]:
    """Normalize scalar/range numbers while preserving the original cell text."""
    raw = normalize_text(value)
    normalized_dash = raw.replace("–", "-").replace("—", "-").replace("－", "-")
    currency = ""
    upper = normalized_dash.upper()
    for token, canonical in sorted(CURRENCY_ALIASES.items(), key=lambda item: len(item[0]), reverse=True):
        if token.upper() in upper:
            currency = canonical
            break
    tax_text = f"{field_name} {raw}".casefold()
    tax_status = "tax_included" if "含税" in tax_text else "tax_excluded" if any(token in tax_text for token in ("未税", "不含税")) else ""

    range_match = RANGE_RE.match(normalized_dash.replace(",", ""))
    if range_match:
        unit_raw = range_match.group("unit").upper()
        return {
            "numeric_value": None,
            "numeric_min": float(range_match.group("minimum")),
            "numeric_max": float(range_match.group("maximum")),
            "unit": UNIT_ALIASES.get(unit_raw, unit_raw),
            "currency": currency or CURRENCY_ALIASES.get((range_match.group("prefix") or "").upper(), ""),
            "tax_status": tax_status,
            "quantity_basis": normalize_text(field_name) if field_key in {"unit_price", "cost_price", "final_price"} else "",
            "normalized_value": normalized_dash,
        }

    numeric_value, unit = parse_numeric(normalized_dash.replace(",", ""))
    canonical_unit = UNIT_ALIASES.get(unit.upper(), unit.upper())
    if canonical_unit in {"¥", "$"}:
        currency = CURRENCY_ALIASES.get(canonical_unit, currency)
        canonical_unit = ""
    return {
        "numeric_value": numeric_value,
        "numeric_min": numeric_value,
        "numeric_max": numeric_value,
        "unit": canonical_unit,
        "currency": currency,
        "tax_status": tax_status,
        "quantity_basis": normalize_text(field_name) if field_key in {"unit_price", "cost_price", "final_price"} else "",
        "normalized_value": normalized_dash,
    }


def model_variants(model: str) -> list[str]:
    normalized = normalize_text(model).upper().replace("_", "-").replace(" ", "")
    return list(dict.fromkeys(item for item in (normalized, normalized.replace("-", "")) if item))


def stable_row_id(document_id: str, sheet_name: str, row_number: int) -> int:
    payload = f"{document_id}\0{sheet_name}\0{row_number}".encode("utf-8")
    return int.from_bytes(hashlib.sha256(payload).digest()[:7], "big")


def create_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        PRAGMA foreign_keys = ON;
        CREATE TABLE index_metadata (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        CREATE TABLE documents (
            document_id TEXT PRIMARY KEY,
            kb_id TEXT NOT NULL,
            kb_key TEXT NOT NULL,
            document_name TEXT NOT NULL,
            storage_bucket TEXT NOT NULL,
            storage_name TEXT NOT NULL,
            content_hash TEXT NOT NULL,
            size_bytes INTEGER NOT NULL,
            indexed_at TEXT NOT NULL,
            sheet_count INTEGER NOT NULL DEFAULT 0,
            row_count INTEGER NOT NULL DEFAULT 0,
            model_count INTEGER NOT NULL DEFAULT 0
            ,source_hash TEXT NOT NULL DEFAULT ''
            ,hash_algorithm TEXT NOT NULL DEFAULT 'sha256'
            ,source_nas_id TEXT NOT NULL DEFAULT ''
            ,source_path TEXT NOT NULL DEFAULT ''
            ,source_version TEXT NOT NULL DEFAULT ''
            ,effective_status TEXT NOT NULL DEFAULT ''
            ,document_type TEXT NOT NULL DEFAULT ''
            ,year TEXT NOT NULL DEFAULT ''
            ,month TEXT NOT NULL DEFAULT ''
            ,season TEXT NOT NULL DEFAULT ''
            ,business_version TEXT NOT NULL DEFAULT ''
            ,authority_technical INTEGER NOT NULL DEFAULT 0
            ,authority_price INTEGER NOT NULL DEFAULT 0
            ,authority_sales_cost INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE document_aliases (
            document_id TEXT PRIMARY KEY,
            canonical_document_id TEXT NOT NULL REFERENCES documents(document_id) ON DELETE CASCADE,
            kb_id TEXT NOT NULL,
            document_name TEXT NOT NULL,
            source_path TEXT NOT NULL DEFAULT '',
            source_version TEXT NOT NULL DEFAULT '',
            effective_status TEXT NOT NULL DEFAULT '',
            content_hash TEXT NOT NULL
        );
        CREATE TABLE sheets (
            document_id TEXT NOT NULL REFERENCES documents(document_id) ON DELETE CASCADE,
            sheet_name TEXT NOT NULL,
            sheet_index INTEGER NOT NULL,
            header_rows_json TEXT NOT NULL,
            row_count INTEGER NOT NULL,
            column_count INTEGER NOT NULL,
            PRIMARY KEY (document_id, sheet_name)
        );
        CREATE TABLE rows (
            id INTEGER PRIMARY KEY,
            document_id TEXT NOT NULL REFERENCES documents(document_id) ON DELETE CASCADE,
            sheet_name TEXT NOT NULL,
            row_number INTEGER NOT NULL,
            header_row_number INTEGER,
            section TEXT NOT NULL DEFAULT '',
            model_tokens TEXT NOT NULL DEFAULT '',
            row_text TEXT NOT NULL,
            values_json TEXT NOT NULL,
            row_hash TEXT NOT NULL,
            UNIQUE (document_id, sheet_name, row_number)
        );
        CREATE TABLE row_fields (
            row_id INTEGER NOT NULL REFERENCES rows(id) ON DELETE CASCADE,
            column_index INTEGER NOT NULL,
            field_name TEXT NOT NULL,
            field_key TEXT NOT NULL,
            raw_value TEXT NOT NULL,
            numeric_value REAL,
            unit TEXT NOT NULL DEFAULT '',
            inherited INTEGER NOT NULL DEFAULT 0,
            numeric_min REAL,
            numeric_max REAL,
            currency TEXT NOT NULL DEFAULT '',
            tax_status TEXT NOT NULL DEFAULT '',
            quantity_basis TEXT NOT NULL DEFAULT '',
            normalized_value TEXT NOT NULL DEFAULT '',
            PRIMARY KEY (row_id, column_index)
        );
        CREATE TABLE row_models (
            model TEXT NOT NULL,
            row_id INTEGER NOT NULL REFERENCES rows(id) ON DELETE CASCADE,
            document_id TEXT NOT NULL,
            sheet_name TEXT NOT NULL,
            row_number INTEGER NOT NULL,
            PRIMARY KEY (model, row_id)
        );
        CREATE INDEX idx_row_models_model ON row_models(model);
        CREATE INDEX idx_row_models_document ON row_models(document_id);
        CREATE INDEX idx_rows_document_sheet ON rows(document_id, sheet_name, row_number);
        CREATE INDEX idx_row_fields_key ON row_fields(field_key);
        CREATE INDEX idx_documents_content_hash ON documents(kb_id, content_hash);
        CREATE INDEX idx_documents_effective ON documents(kb_id, effective_status);
        CREATE VIRTUAL TABLE rows_fts USING fts5(
            row_id UNINDEXED,
            document_name,
            sheet_name,
            model_tokens,
            row_text,
            tokenize='unicode61 remove_diacritics 2'
        );
        """
    )


def _trim(values: list[str]) -> list[str]:
    while values and not values[-1]:
        values.pop()
    return values


def _header_names(values: list[str]) -> list[str]:
    names = []
    seen: dict[str, int] = {}
    for column_index, value in enumerate(values, start=1):
        base = normalize_text(value) or f"column_{column_index}"
        seen[base] = seen.get(base, 0) + 1
        names.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return names


def index_workbook_blob(
    connection: sqlite3.Connection,
    document: dict[str, Any],
    blob: bytes,
    max_columns: int = 256,
) -> dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(blob), read_only=True, data_only=True, keep_links=False)
    document_id = str(document["document_id"])
    document_name = str(document["document_name"])
    indexed_at = datetime.now().astimezone().isoformat(timespec="seconds")
    connection.execute(
        """
        INSERT INTO documents (
            document_id, kb_id, kb_key, document_name, storage_bucket, storage_name,
            content_hash, size_bytes, indexed_at, source_hash, hash_algorithm,
            source_nas_id, source_path, source_version, effective_status,
            document_type, year, month, season, business_version,
            authority_technical, authority_price, authority_sales_cost
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            document_id,
            document["kb_id"],
            document["kb_key"],
            document_name,
            document["storage_bucket"],
            document["storage_name"],
            document.get("content_hash", ""),
            int(document.get("size_bytes", len(blob))),
            indexed_at,
            document.get("source_hash", document.get("content_hash", "")),
            document.get("hash_algorithm", "sha256"),
            document.get("source_nas_id", ""),
            document.get("source_path", ""),
            document.get("source_version", ""),
            document.get("effective_status", ""),
            document.get("document_type", ""),
            document.get("year", ""),
            document.get("month", ""),
            document.get("season", ""),
            document.get("business_version", ""),
            int(document.get("authority_technical", 0) or 0),
            int(document.get("authority_price", 0) or 0),
            int(document.get("authority_sales_cost", 0) or 0),
        ),
    )

    total_rows = 0
    all_models: set[str] = set()
    sheet_reports = []
    try:
        for sheet_index, worksheet in enumerate(workbook.worksheets):
            bounded_columns = max(1, min(int(worksheet.max_column or 1), max_columns))
            current_headers: list[str] | None = None
            current_field_keys: list[str] | None = None
            current_header_row: int | None = None
            current_section = ""
            inherited_models: list[str] = []
            header_rows = []
            sheet_rows = 0
            observed_columns = 0
            previous_values: list[str] = []
            truncated_columns = False

            for row_number, cells in enumerate(
                worksheet.iter_rows(max_col=bounded_columns, values_only=True), start=1
            ):
                values = _trim([normalize_text(cell) for cell in cells])
                if not any(values):
                    continue
                observed_columns = max(observed_columns, len(values))
                if worksheet.max_column and worksheet.max_column > max_columns:
                    truncated_columns = True
                detected_keys = detect_header(values)
                header_values = values
                if detected_keys is None and not current_headers and previous_values:
                    combined = [
                        " ".join(item for item in (previous_values[index] if index < len(previous_values) else "", value) if item)
                        for index, value in enumerate(values)
                    ]
                    combined_keys = detect_header(combined)
                    if combined_keys is not None:
                        detected_keys = combined_keys
                        header_values = combined
                if detected_keys is not None:
                    current_headers = _header_names(header_values)
                    current_field_keys = detected_keys
                    current_header_row = row_number if header_values is values else max(1, row_number - 1)
                    current_section = ""
                    inherited_models = []
                    header_rows.append(
                        {"row": row_number, "headers": current_headers, "field_keys": current_field_keys}
                    )
                    previous_values = []
                    continue
                if current_headers is None or current_field_keys is None:
                    previous_values = values
                    continue

                nonempty = [(index, value) for index, value in enumerate(values) if value]
                model_columns = [
                    index for index, key in enumerate(current_field_keys) if key in MODEL_FIELD_KEYS
                ]
                explicit_models = []
                for index in model_columns:
                    if index < len(values) and values[index]:
                        explicit_models.extend(extract_row_models(values[index]))
                explicit_models = list(dict.fromkeys(explicit_models))

                if len(nonempty) <= 1 and not explicit_models:
                    current_section = nonempty[0][1]
                    inherited_models = []
                    continue
                if explicit_models:
                    inherited_models = explicit_models
                row_models = explicit_models or inherited_models

                fields = []
                row_parts = []
                for column_index in range(max(len(values), len(current_headers))):
                    raw_value = values[column_index] if column_index < len(values) else ""
                    field_name = (
                        current_headers[column_index]
                        if column_index < len(current_headers)
                        else f"column_{column_index + 1}"
                    )
                    field_key = (
                        current_field_keys[column_index]
                        if column_index < len(current_field_keys)
                        else ""
                    )
                    inherited = 0
                    if not raw_value and field_key == "model" and row_models:
                        raw_value = ", ".join(row_models)
                        inherited = 1
                    if not raw_value:
                        continue
                    normalized_value = parse_business_value(raw_value, field_name, field_key)
                    fields.append(
                        {
                            "column_index": column_index + 1,
                            "field_name": field_name,
                            "field_key": field_key,
                            "raw_value": raw_value,
                            **normalized_value,
                            "inherited": inherited,
                        }
                    )
                    inherited_note = " (继承)" if inherited else ""
                    row_parts.append(f"{field_name}{inherited_note}: {raw_value}")
                if not fields:
                    continue

                row_text_parts = [f"文件: {document_name}", f"工作表: {worksheet.title}"]
                if current_section:
                    row_text_parts.append(f"分类: {current_section}")
                row_text_parts.extend(row_parts)
                row_text = " | ".join(row_text_parts)
                values_json = json.dumps(
                    {"section": current_section, "models": row_models, "fields": fields},
                    ensure_ascii=False,
                    separators=(",", ":"),
                )
                row_hash = hashlib.sha256(row_text.encode("utf-8")).hexdigest()
                row_id = stable_row_id(document_id, worksheet.title, row_number)
                connection.execute(
                    """
                    INSERT INTO rows (
                        id, document_id, sheet_name, row_number, header_row_number, section,
                        model_tokens, row_text, values_json, row_hash
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        row_id,
                        document_id,
                        worksheet.title,
                        row_number,
                        current_header_row,
                        current_section,
                        " ".join(row_models),
                        row_text,
                        values_json,
                        row_hash,
                    ),
                )
                connection.executemany(
                    """
                        INSERT INTO row_fields (
                            row_id, column_index, field_name, field_key, raw_value,
                            numeric_value, unit, inherited, numeric_min, numeric_max,
                            currency, tax_status, quantity_basis, normalized_value
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    [
                        (
                            row_id,
                            field["column_index"],
                            field["field_name"],
                            field["field_key"],
                            field["raw_value"],
                            field["numeric_value"],
                            field["unit"],
                            field["inherited"],
                            field["numeric_min"],
                            field["numeric_max"],
                            field["currency"],
                            field["tax_status"],
                            field["quantity_basis"],
                            field["normalized_value"],
                        )
                        for field in fields
                    ],
                )
                for model in row_models:
                    for model_variant in model_variants(model):
                        connection.execute(
                            """
                            INSERT OR IGNORE INTO row_models
                                (model, row_id, document_id, sheet_name, row_number)
                            VALUES (?, ?, ?, ?, ?)
                            """,
                            (model_variant, row_id, document_id, worksheet.title, row_number),
                        )
                    all_models.add(model)
                connection.execute(
                    """
                    INSERT INTO rows_fts
                        (row_id, document_name, sheet_name, model_tokens, row_text)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (row_id, document_name, worksheet.title, " ".join(row_models), row_text),
                )
                sheet_rows += 1
                total_rows += 1

            connection.execute(
                """
                INSERT INTO sheets (
                    document_id, sheet_name, sheet_index, header_rows_json,
                    row_count, column_count
                ) VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    document_id,
                    worksheet.title,
                    sheet_index,
                    json.dumps(header_rows, ensure_ascii=False, separators=(",", ":")),
                    sheet_rows,
                    observed_columns,
                ),
            )
            sheet_reports.append(
                {
                    "name": worksheet.title,
                    "rows": sheet_rows,
                    "columns": observed_columns,
                    "header_rows": [item["row"] for item in header_rows],
                    "truncated_columns": truncated_columns,
                }
            )
    finally:
        workbook.close()

    connection.execute(
        """
        UPDATE documents
        SET sheet_count = ?, row_count = ?, model_count = ?
        WHERE document_id = ?
        """,
        (len(workbook.sheetnames), total_rows, len(all_models), document_id),
    )
    return {
        "document_id": document_id,
        "document_name": document_name,
        "sheets": len(workbook.sheetnames),
        "rows": total_rows,
        "models": len(all_models),
        "sheet_details": sheet_reports,
    }


def database_metadata(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    try:
        with sqlite3.connect(f"file:{path}?mode=ro", uri=True) as connection:
            return dict(connection.execute("SELECT key, value FROM index_metadata"))
    except (OSError, sqlite3.Error):
        return {}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return normalize_text(value)


def _knowledge_base_specs(config: dict[str, Any]) -> list[dict[str, str]]:
    specs = config.get("knowledge_bases") or []
    if not specs:
        specs = [
            {
                "key": item["knowledge_base_key"],
                "name": item["knowledge_base_name"],
            }
            for item in config.get("documents", [])
        ]
    result = []
    seen = set()
    for item in specs:
        key = str(item.get("key") or item.get("knowledge_base_key") or "").strip()
        name = str(item.get("name") or item.get("knowledge_base_name") or "").strip()
        if not key or not name or (key, name) in seen:
            continue
        seen.add((key, name))
        result.append({"key": key, "name": name})
    return result


def resolve_documents(config: dict[str, Any]) -> list[dict[str, Any]]:
    """Discover parsed Excel documents and attach their current business metadata."""
    from api.db.db_models import Document
    from api.db.services.doc_metadata_service import DocMetadataService
    from api.db.services.file2document_service import File2DocumentService
    from api.db.services.knowledgebase_service import KnowledgebaseService
    from common import settings
    from common.constants import TaskStatus
    from ragflow_env import get_target_tenant_id

    tenant_id = get_target_tenant_id()
    explicit = {str(item["document_id"]): item for item in config.get("documents", [])}
    auto_discover = bool(config.get("auto_discover", True))
    suffixes = {
        str(item).lower().lstrip(".")
        for item in config.get("supported_suffixes", ["xlsx", "xlsm"])
    }
    max_file_bytes = int(config.get("max_file_bytes", 150 * 1024 * 1024))
    max_documents = max(1, int(config.get("max_documents", 10000)))
    resolved = []
    for spec in _knowledge_base_specs(config):
        ok, knowledge_base = KnowledgebaseService.get_by_name(spec["name"], tenant_id)
        if not ok:
            raise RuntimeError(f"knowledge base not found: {spec['name']}")
        query = Document.select().where(
            (Document.kb_id == knowledge_base.id)
            & (Document.status == "1")
            & (Document.run == TaskStatus.DONE.value)
            & (Document.chunk_num > 0)
        )
        metadata = DocMetadataService.get_metadata_for_documents(None, knowledge_base.id)
        for document in query:
            document_id = str(document.id)
            if not auto_discover and document_id not in explicit:
                continue
            suffix = str(document.suffix or Path(str(document.name or "")).suffix).lower().lstrip(".")
            if suffix not in suffixes or int(document.size or 0) <= 0 or int(document.size or 0) > max_file_bytes:
                continue
            bucket, storage_name = File2DocumentService.get_storage_address(doc_id=document.id)
            if not settings.STORAGE_IMPL.obj_exist(bucket, storage_name):
                continue
            meta = dict(metadata.get(document_id) or {})
            source_path = _text(meta.get("nas_relative_path") or document.location or document.name)
            inferred = infer_business_metadata(source_path or str(document.name), kb_key=spec["key"])
            merged_meta = {**inferred, **meta}
            resolved.append(
                {
                **explicit.get(document_id, {}),
                "document_id": str(document.id),
                "document_name": str(document.name),
                "kb_id": str(document.kb_id),
                "kb_key": spec["key"],
                "storage_bucket": str(bucket),
                "storage_name": str(storage_name),
                "source_hash": str(document.content_hash or ""),
                "content_hash": "",
                "size_bytes": int(document.size or 0),
                "source_nas_id": _text(merged_meta.get("source_nas_id")),
                "source_path": source_path,
                "source_version": _text(merged_meta.get("source_version")),
                "effective_status": _text(merged_meta.get("effective_status") or "active"),
                "document_type": _text(merged_meta.get("document_type")),
                "year": _text(merged_meta.get("year") or merged_meta.get("doc_year")),
                "month": _text(merged_meta.get("month")),
                "season": _text(merged_meta.get("season")),
                "business_version": _text(merged_meta.get("business_version")),
                "authority_technical": int(merged_meta.get("authority_technical") or 0),
                "authority_price": int(merged_meta.get("authority_price") or 0),
                "authority_sales_cost": int(merged_meta.get("authority_sales_cost") or 0),
                "priority": document_id in explicit,
                }
            )
    resolved.sort(
        key=lambda item: (
            not item["priority"],
            str(item.get("effective_status") or "").casefold() in {"historical", "inactive", "expired"},
            item["kb_key"],
            item["document_name"],
            item["document_id"],
        )
    )
    return resolved[:max_documents]


def _source_fingerprint(document: dict[str, Any]) -> str:
    payload = {
        "document_id": document["document_id"],
        "source_hash": document.get("source_hash", ""),
        "size_bytes": document["size_bytes"],
        "storage_bucket": document["storage_bucket"],
        "storage_name": document["storage_name"],
    }
    return hashlib.sha256(json.dumps(payload, sort_keys=True, separators=(",", ":")).encode()).hexdigest()


def _metadata_signature(document: dict[str, Any]) -> str:
    keys = (
        "document_id", "document_name", "source_nas_id", "source_path", "source_version",
        "effective_status", "document_type", "year", "month", "season", "business_version",
        "authority_technical", "authority_price", "authority_sales_cost",
    )
    payload = {key: document.get(key, "") for key in keys}
    return hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode()).hexdigest()


def _load_manifest(path: Path) -> dict[str, Any]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return data if data.get("schema_version") == SCHEMA_VERSION else {}
    except (OSError, json.JSONDecodeError):
        return {}


def _materialize_hashes(
    documents: list[dict[str, Any]], previous: dict[str, Any]
) -> tuple[list[dict[str, Any]], dict[str, bytes], list[dict[str, str]]]:
    from common import settings

    previous_documents = previous.get("documents", {})
    blob_cache: dict[str, bytes] = {}
    hash_cache: dict[tuple[str, str, int], tuple[str, bytes | None]] = {}
    ready = []
    errors = []
    for document in documents:
        fingerprint = _source_fingerprint(document)
        old = previous_documents.get(document["document_id"], {})
        actual_hash = ""
        if old.get("source_fingerprint") == fingerprint and old.get("content_hash"):
            actual_hash = str(old["content_hash"])
        cache_key = (document["kb_id"], document.get("source_hash", ""), document["size_bytes"])
        if not actual_hash and document.get("source_hash") and cache_key in hash_cache:
            actual_hash, cached_blob = hash_cache[cache_key]
            if cached_blob is not None:
                blob_cache[document["document_id"]] = cached_blob
        if not actual_hash:
            try:
                blob = settings.STORAGE_IMPL.get(document["storage_bucket"], document["storage_name"])
                actual_hash = hashlib.sha256(blob).hexdigest()
                blob_cache[document["document_id"]] = blob
                if document.get("source_hash"):
                    hash_cache[cache_key] = (actual_hash, blob)
            except Exception as exc:
                errors.append({"document_id": document["document_id"], "stage": "hash", "error": repr(exc)})
                continue
        ready.append({**document, "content_hash": actual_hash, "hash_algorithm": "sha256", "source_fingerprint": fingerprint})
    return ready, blob_cache, errors


def _read_shard_summary(path: Path) -> dict[str, Any]:
    with sqlite3.connect(f"file:{path}?mode=ro", uri=True) as connection:
        row = connection.execute(
            "SELECT document_id, document_name, sheet_count, row_count, model_count FROM documents"
        ).fetchone()
        return {
            "document_id": row[0], "document_name": row[1], "sheets": row[2],
            "rows": row[3], "models": row[4], "shard": str(path),
        }


def _build_or_reuse_shard(
    document: dict[str, Any], shard_path: Path, blob_cache: dict[str, bytes],
    max_columns: int, force: bool,
) -> tuple[dict[str, Any], bool]:
    metadata = database_metadata(shard_path)
    metadata_signature = _metadata_signature(document)
    reusable = (
        not force
        and metadata.get("schema_version") == SCHEMA_VERSION
        and metadata.get("content_hash") == document["content_hash"]
        and metadata.get("document_id") == document["document_id"]
        and metadata.get("metadata_signature") == metadata_signature
    )
    if reusable:
        return _read_shard_summary(shard_path), True

    from common import settings

    blob = blob_cache.get(document["document_id"])
    if blob is None:
        blob = settings.STORAGE_IMPL.get(document["storage_bucket"], document["storage_name"])
    if hashlib.sha256(blob).hexdigest() != document["content_hash"]:
        raise RuntimeError(f"content changed while indexing: {document['document_id']}")
    shard_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = tempfile.NamedTemporaryFile(prefix=f".{shard_path.name}.", suffix=".tmp", dir=shard_path.parent, delete=False)
    temp_path = Path(temporary.name)
    temporary.close()
    try:
        with sqlite3.connect(temp_path) as connection:
            create_schema(connection)
            report = index_workbook_blob(connection, document, blob, max_columns=max_columns)
            values = {
                "schema_version": SCHEMA_VERSION,
                "content_hash": document["content_hash"],
                "document_id": document["document_id"],
                "metadata_signature": metadata_signature,
                "built_at": datetime.now().astimezone().isoformat(timespec="seconds"),
            }
            connection.executemany("INSERT INTO index_metadata (key, value) VALUES (?, ?)", values.items())
            if connection.execute("PRAGMA integrity_check").fetchone()[0] != "ok":
                raise RuntimeError(f"SQLite integrity check failed: {document['document_id']}")
            connection.commit()
        os.replace(temp_path, shard_path)
        os.chmod(shard_path, 0o644)
        return {**report, "shard": str(shard_path)}, False
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _merge_shard(connection: sqlite3.Connection, path: Path) -> None:
    connection.execute("ATTACH DATABASE ? AS source_shard", (str(path),))
    try:
        for table in ("documents", "sheets", "rows", "row_fields", "row_models"):
            connection.execute(f"INSERT INTO {table} SELECT * FROM source_shard.{table}")
        connection.execute(
            "INSERT INTO rows_fts (row_id, document_name, sheet_name, model_tokens, row_text) "
            "SELECT row_id, document_name, sheet_name, model_tokens, row_text FROM source_shard.rows_fts"
        )
    finally:
        connection.commit()
        connection.execute("DETACH DATABASE source_shard")


def write_report(path: Path, report: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")


def build_index(config_path: Path, database_path: Path, report_path: Path, force: bool = False) -> dict:
    from common import settings

    started_at = datetime.now().astimezone().isoformat(timespec="seconds")
    config = json.loads(config_path.read_text(encoding="utf-8"))
    settings.init_settings()
    discovered = resolve_documents(config)
    if not discovered:
        raise RuntimeError("no parsed Excel documents discovered")

    database_path.parent.mkdir(parents=True, exist_ok=True)
    shards_dir = database_path.parent / f"{database_path.stem}_shards"
    shards_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = shards_dir / "manifest.json"
    previous_manifest = _load_manifest(manifest_path)
    documents, blob_cache, errors = _materialize_hashes(discovered, previous_manifest)
    priority_ids = {item["document_id"] for item in discovered if item.get("priority")}
    priority_errors = [item for item in errors if item["document_id"] in priority_ids]
    if priority_errors:
        raise RuntimeError(f"priority Excel hashing failed: {priority_errors}")

    grouped: dict[str, list[dict[str, Any]]] = {}
    for document in documents:
        group_key = f"{document['kb_id']}:{document['content_hash']}"
        grouped.setdefault(group_key, []).append(document)

    previous_groups = previous_manifest.get("groups", {})
    successful_groups = []
    document_reports = []
    reused_shards = 0
    rebuilt_shards = 0
    for group_key, members in sorted(grouped.items()):
        members.sort(
            key=lambda item: (
                not item.get("priority", False),
                str(item.get("effective_status") or "").casefold() in {"historical", "inactive", "expired", "missing_from_source"},
                item["document_name"],
                item["document_id"],
            )
        )
        previous_canonical = str(previous_groups.get(group_key, {}).get("canonical_document_id") or "")
        canonical = next((item for item in members if item["document_id"] == previous_canonical), members[0])
        shard_path = shards_dir / f"{canonical['kb_id']}_{canonical['content_hash']}.sqlite3"
        try:
            shard_report, reused = _build_or_reuse_shard(
                canonical,
                shard_path,
                blob_cache,
                max_columns=int(config.get("max_columns", 256)),
                force=force,
            )
        except Exception as exc:
            failure = {
                "document_id": canonical["document_id"],
                "stage": "parse",
                "error": repr(exc),
            }
            errors.append(failure)
            if any(item.get("priority") for item in members):
                raise RuntimeError(f"priority Excel indexing failed: {failure}") from exc
            continue
        reused_shards += int(reused)
        rebuilt_shards += int(not reused)
        document_reports.append({**shard_report, "aliases": len(members), "reused": reused})
        successful_groups.append(
            {
                "group_key": group_key,
                "canonical": canonical,
                "members": members,
                "shard_path": shard_path,
            }
        )

    if not successful_groups:
        raise RuntimeError("all discovered Excel documents failed to index")

    signature_payload = {
        "schema_version": SCHEMA_VERSION,
        "max_columns": int(config.get("max_columns", 256)),
        "groups": [
            {
                "key": item["group_key"],
                "canonical": item["canonical"]["document_id"],
                "metadata": _metadata_signature(item["canonical"]),
                "aliases": [
                    {
                        "document_id": member["document_id"],
                        "source_path": member.get("source_path", ""),
                        "source_version": member.get("source_version", ""),
                        "effective_status": member.get("effective_status", ""),
                    }
                    for member in item["members"]
                ],
            }
            for item in successful_groups
        ],
    }
    signature = hashlib.sha256(
        json.dumps(signature_payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode()
    ).hexdigest()
    previous_database = database_metadata(database_path)
    unchanged = (
        not force
        and previous_database.get("schema_version") == SCHEMA_VERSION
        and previous_database.get("source_signature") == signature
    )

    if not unchanged:
        temp_file = tempfile.NamedTemporaryFile(
            prefix=f".{database_path.name}.", suffix=".tmp", dir=database_path.parent, delete=False
        )
        temp_path = Path(temp_file.name)
        temp_file.close()
        try:
            with sqlite3.connect(temp_path) as connection:
                create_schema(connection)
                for item in successful_groups:
                    _merge_shard(connection, item["shard_path"])
                    canonical_id = item["canonical"]["document_id"]
                    connection.executemany(
                        """
                        INSERT INTO document_aliases (
                            document_id, canonical_document_id, kb_id, document_name,
                            source_path, source_version, effective_status, content_hash
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        [
                            (
                                member["document_id"], canonical_id, member["kb_id"],
                                member["document_name"], member.get("source_path", ""),
                                member.get("source_version", ""), member.get("effective_status", ""),
                                member["content_hash"],
                            )
                            for member in item["members"]
                        ],
                    )
                finished_at = datetime.now().astimezone().isoformat(timespec="seconds")
                metadata = {
                    "schema_version": SCHEMA_VERSION,
                    "source_signature": signature,
                    "built_at": finished_at,
                    "config_path": str(config_path),
                    "discovered_documents": str(len(discovered)),
                    "canonical_documents": str(len(successful_groups)),
                }
                connection.executemany("INSERT INTO index_metadata (key, value) VALUES (?, ?)", metadata.items())
                connection.execute("ANALYZE")
                integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
                if integrity != "ok":
                    raise RuntimeError(f"SQLite integrity check failed: {integrity}")
                connection.commit()
            os.replace(temp_path, database_path)
            os.chmod(database_path, 0o644)
        except Exception:
            temp_path.unlink(missing_ok=True)
            raise
    else:
        finished_at = datetime.now().astimezone().isoformat(timespec="seconds")

    next_manifest = {
        "schema_version": SCHEMA_VERSION,
        "built_at": finished_at,
        "source_signature": signature,
        "documents": {
            member["document_id"]: {
                "source_fingerprint": member["source_fingerprint"],
                "content_hash": member["content_hash"],
                "group_key": item["group_key"],
                "canonical_document_id": item["canonical"]["document_id"],
            }
            for item in successful_groups
            for member in item["members"]
        },
        "groups": {
            item["group_key"]: {
                "canonical_document_id": item["canonical"]["document_id"],
                "shard": item["shard_path"].name,
                "aliases": [member["document_id"] for member in item["members"]],
            }
            for item in successful_groups
        },
    }
    manifest_temp = manifest_path.with_suffix(".json.tmp")
    manifest_temp.write_text(json.dumps(next_manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(manifest_temp, manifest_path)
    referenced_shards = {item["shard_path"].resolve() for item in successful_groups}
    for old_shard in shards_dir.glob("*.sqlite3"):
        if old_shard.resolve() not in referenced_shards:
            old_shard.unlink(missing_ok=True)

    report = {
        "status": "unchanged" if unchanged else "built",
        "started_at": started_at,
        "finished_at": finished_at,
        "database_path": str(database_path),
        "manifest_path": str(manifest_path),
        "source_signature": signature,
        "documents": document_reports,
        "errors": errors,
        "totals": {
            "discovered_documents": len(discovered),
            "canonical_documents": len(successful_groups),
            "duplicate_aliases": sum(max(0, len(item["members"]) - 1) for item in successful_groups),
            "reused_shards": reused_shards,
            "rebuilt_shards": rebuilt_shards,
            "sheets": sum(item["sheets"] for item in document_reports),
            "rows": sum(item["rows"] for item in document_reports),
            "models": sum(item["models"] for item in document_reports),
        },
    }
    write_report(report_path, report)
    return report


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG_PATH)
    parser.add_argument("--database", type=Path, default=DEFAULT_DATABASE_PATH)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT_PATH)
    parser.add_argument("--force", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    report = build_index(args.config, args.database, args.report, force=args.force)
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
