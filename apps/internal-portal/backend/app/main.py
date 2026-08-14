from __future__ import annotations

import asyncio
import ipaddress
import html as html_lib
from io import BytesIO
import json
import logging
import os
import secrets
import time
from typing import Any, AsyncIterator
from urllib.parse import quote

import httpx
from openpyxl import load_workbook
from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import JSONResponse, RedirectResponse, Response, StreamingResponse
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer

from .document_headers import _document_response_headers
from .history_store import HistoryStore
from .ops_status import evaluation_status, sync_status


logging.basicConfig(level=os.getenv("LOG_LEVEL", "INFO"))
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)
logger = logging.getLogger("waimao-internal-portal")

app = FastAPI(title="Waimao Internal Portal API", docs_url=None, redoc_url=None)


def _secret_file_values() -> dict[str, str]:
    values: dict[str, str] = {}
    path = os.getenv("INTERNAL_PORTAL_SECRET_FILE", "/app/secrets/internal-portal.env")
    try:
        with open(path, encoding="utf-8") as secret_file:
            for line in secret_file:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                values[key.strip()] = value.strip().strip('"').strip("'")
    except FileNotFoundError:
        pass
    return values


SECRET_FILE_VALUES = _secret_file_values()


def _setting(name: str, default: str = "") -> str:
    return os.getenv(name, "").strip() or SECRET_FILE_VALUES.get(name, default).strip()


RAGFLOW_BASE_URL = os.getenv("RAGFLOW_BASE_URL", "http://ragflow-cpu:9380").rstrip("/")
RAGFLOW_API_TOKEN = _setting("RAGFLOW_API_TOKEN")
AUTH_MODE = _setting("AUTH_MODE", "trusted_lan").lower()
SESSION_COOKIE = "letouch_internal_session"
SESSION_MAX_AGE = int(os.getenv("SESSION_MAX_AGE", "28800"))
SESSION_SECRET = _setting("SESSION_SECRET", "change-this-before-production")
COOKIE_SECURE = _setting("COOKIE_SECURE", "0") == "1"
WECOM_CORP_ID = _setting("WECOM_CORP_ID")
WECOM_AGENT_ID = _setting("WECOM_AGENT_ID")
WECOM_SECRET = _setting("WECOM_SECRET")
WECOM_REDIRECT_URI = _setting("WECOM_REDIRECT_URI")
SYNC_STATUS_ROOT = os.getenv("SYNC_STATUS_ROOT", "/app/ops/incremental_sync")
EVALUATION_STATUS_ROOT = os.getenv("EVALUATION_STATUS_ROOT", "/app/ops/evaluation")
HISTORY_DB_PATH = os.getenv("HISTORY_DB_PATH", "/app/data/conversations.sqlite3")
HISTORY_ACTOR_COOKIE = "letouch_internal_actor"
HISTORY_ACTOR_MAX_AGE = 60 * 60 * 24 * 365
serializer = URLSafeTimedSerializer(SESSION_SECRET, salt="letouch-internal-portal")
history_store = HistoryStore(HISTORY_DB_PATH)

_wecom_token = ""
_wecom_token_expires_at = 0.0
_wecom_token_lock = asyncio.Lock()

ASSISTANT_CHAT_IDS = {
    "purchase": "<configure-purchase-chat-id>",
    "sales": "<configure-sales-chat-id>",
    "product": "<configure-product-chat-id>",
}


def _networks() -> list[ipaddress._BaseNetwork]:
    configured = os.getenv(
        "TRUSTED_NETWORKS",
        "127.0.0.1/32,::1/128,192.0.2.0/24,198.51.100.0/24,203.0.113.0/24",
    )
    networks: list[ipaddress._BaseNetwork] = []
    for item in configured.split(","):
        try:
            networks.append(ipaddress.ip_network(item.strip(), strict=False))
        except ValueError:
            logger.warning("Ignoring invalid TRUSTED_NETWORKS entry: %s", item)
    return networks


TRUSTED_NETWORKS = _networks()


def _client_ip(request: Request) -> ipaddress._BaseAddress:
    forwarded = request.headers.get("x-real-ip") or request.client.host
    try:
        return ipaddress.ip_address(forwarded)
    except ValueError as exc:
        raise HTTPException(status_code=403, detail="无法确认访问来源") from exc


def _in_trusted_network(request: Request) -> bool:
    address = _client_ip(request)
    return any(address in network for network in TRUSTED_NETWORKS)


def _session_payload(request: Request) -> dict[str, Any] | None:
    value = request.cookies.get(SESSION_COOKIE)
    if not value:
        return None
    try:
        payload = serializer.loads(value, max_age=SESSION_MAX_AGE)
    except (BadSignature, SignatureExpired):
        return None
    return payload if isinstance(payload, dict) else None


def _safe_next_path(value: str) -> str:
    if value.startswith("/internal/") and not value.startswith("//"):
        return value
    return "/internal/"


def _wecom_login_flow(request: Request) -> str:
    user_agent = request.headers.get("user-agent", "").lower()
    return "in_app" if "wxwork" in user_agent else "qr"


async def _wecom_access_token() -> str:
    global _wecom_token, _wecom_token_expires_at

    if _wecom_token and time.monotonic() < _wecom_token_expires_at:
        return _wecom_token

    async with _wecom_token_lock:
        if _wecom_token and time.monotonic() < _wecom_token_expires_at:
            return _wecom_token
        if not WECOM_CORP_ID or not WECOM_SECRET:
            raise HTTPException(status_code=503, detail="企业微信参数尚未配置")

        async with httpx.AsyncClient(timeout=15) as client:
            response = await client.get(
                "https://qyapi.weixin.qq.com/cgi-bin/gettoken",
                params={"corpid": WECOM_CORP_ID, "corpsecret": WECOM_SECRET},
            )
        response.raise_for_status()
        data = response.json()
        if data.get("errcode", 0) != 0 or not data.get("access_token"):
            logger.warning("WeCom token request failed errcode=%s", data.get("errcode"))
            raise HTTPException(status_code=502, detail="企业微信 access_token 获取失败")

        expires_in = max(int(data.get("expires_in", 7200)) - 120, 60)
        _wecom_token = str(data["access_token"])
        _wecom_token_expires_at = time.monotonic() + expires_in
        return _wecom_token


async def _wecom_user_identity(code: str, flow: str) -> dict[str, str]:
    access_token = await _wecom_access_token()
    endpoint = "user/getuserinfo" if flow == "in_app" else "auth/getuserinfo"
    async with httpx.AsyncClient(timeout=15) as client:
        response = await client.get(
            f"https://qyapi.weixin.qq.com/cgi-bin/{endpoint}",
            params={"access_token": access_token, "code": code},
        )
        response.raise_for_status()
        data = response.json()
        if data.get("errcode", 0) != 0:
            logger.warning("WeCom identity request failed flow=%s errcode=%s", flow, data.get("errcode"))
            raise HTTPException(status_code=401, detail="企业微信用户身份验证失败")

        user_id = data.get("UserId") or data.get("userid")
        if not user_id:
            raise HTTPException(status_code=403, detail="仅允许企业内部成员访问")

        name = str(user_id)
        profile_response = await client.get(
            "https://qyapi.weixin.qq.com/cgi-bin/user/get",
            params={"access_token": access_token, "userid": user_id},
        )
        if profile_response.is_success:
            profile = profile_response.json()
            if profile.get("errcode", 0) == 0 and profile.get("name"):
                name = str(profile["name"])

    return {"user_id": str(user_id), "name": name}


def _require_access(request: Request) -> dict[str, Any]:
    if AUTH_MODE == "trusted_lan":
        if not _in_trusted_network(request):
            raise HTTPException(status_code=403, detail="当前设备不在公司内网允许范围")
        return {"mode": AUTH_MODE, "ip": str(_client_ip(request))}

    if AUTH_MODE == "wecom":
        payload = _session_payload(request)
        if not payload:
            raise HTTPException(status_code=401, detail="请先使用企业微信登录")
        return payload

    raise HTTPException(status_code=503, detail="门户认证模式配置错误")


def _history_owner(request: Request) -> tuple[str, str | None]:
    """Return a stable owner key and a new browser cookie when LAN mode has no identity."""
    access = _require_access(request)
    if AUTH_MODE == "wecom":
        user_id = str(access.get("user_id") or "").strip()
        if not user_id:
            raise HTTPException(status_code=401, detail="企业微信用户身份不可用")
        return f"wecom:{user_id}", None

    actor_cookie = request.cookies.get(HISTORY_ACTOR_COOKIE)
    if actor_cookie:
        try:
            payload = serializer.loads(actor_cookie, max_age=HISTORY_ACTOR_MAX_AGE)
            actor_id = str(payload.get("actor_id") or "").strip() if isinstance(payload, dict) else ""
            if actor_id:
                return f"browser:{actor_id}", None
        except (BadSignature, SignatureExpired):
            pass

    actor_id = secrets.token_urlsafe(24)
    return f"browser:{actor_id}", serializer.dumps({"actor_id": actor_id})


def _set_history_cookie(response: JSONResponse, signed_cookie: str | None) -> JSONResponse:
    if signed_cookie:
        response.set_cookie(
            HISTORY_ACTOR_COOKIE,
            signed_cookie,
            max_age=HISTORY_ACTOR_MAX_AGE,
            httponly=True,
            samesite="lax",
            secure=COOKIE_SECURE,
            path="/",
        )
    return response


def _history_assistant_id(value: Any) -> str:
    assistant_id = str(value or "").strip()
    if assistant_id not in ASSISTANT_CHAT_IDS:
        raise HTTPException(status_code=404, detail="未知的知识助手")
    return assistant_id


def _safe_history_messages(value: Any) -> list[dict[str, Any]]:
    if not isinstance(value, list) or not value or len(value) > 100:
        raise HTTPException(status_code=400, detail="历史消息格式不正确")
    messages: list[dict[str, Any]] = []
    for item in value:
        if not isinstance(item, dict) or item.get("role") not in {"user", "assistant"}:
            raise HTTPException(status_code=400, detail="历史消息格式不正确")
        content = item.get("content")
        if not isinstance(content, str) or len(content) > 20000:
            raise HTTPException(status_code=400, detail="历史消息内容不正确")
        message: dict[str, Any] = {"role": item["role"], "content": content}
        references = item.get("references")
        if isinstance(references, list):
            message["references"] = [
                {
                    key: ref[key]
                    for key in ("id", "doc_id", "document_id", "doc_name", "document_name", "similarity", "page_num", "content_with_weight", "content_ltks", "positions")
                    if key in ref
                }
                for ref in references[:20]
                if isinstance(ref, dict)
            ]
        messages.append(message)
    return messages


def _ragflow_headers() -> dict[str, str]:
    if not RAGFLOW_API_TOKEN:
        raise HTTPException(status_code=503, detail="内部代理尚未配置 RAGFlow 服务密钥")
    token = RAGFLOW_API_TOKEN
    if not token.lower().startswith("bearer "):
        token = f"Bearer {token}"
    return {"Authorization": token}


def _safe_chat_request(body: dict[str, Any], assistant_id: str) -> dict[str, Any]:
    chat_id = ASSISTANT_CHAT_IDS.get(assistant_id)
    if not chat_id:
        raise HTTPException(status_code=404, detail="未知的知识助手")
    messages = body.get("messages")
    if not isinstance(messages, list) or not messages:
        raise HTTPException(status_code=400, detail="消息不能为空")
    safe_messages = []
    for message in messages:
        if not isinstance(message, dict) or message.get("role") not in {"user", "assistant"}:
            raise HTTPException(status_code=400, detail="消息格式不正确")
        content = message.get("content")
        if not isinstance(content, str) or not content.strip() or len(content) > 20000:
            raise HTTPException(status_code=400, detail="消息内容不正确")
        safe_messages.append({"role": message["role"], "content": content})

    return {
        "chat_id": chat_id,
        "session_id": body.get("session_id") or None,
        "messages": safe_messages,
        "pass_all_history_messages": True,
        "reasoning": False,
        "stream": True,
    }


def _spreadsheet_html(data: bytes, filename: str) -> str:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[str] = []
    for row_index, row in enumerate(sheet.iter_rows(max_row=500, max_col=50), start=1):
        cells = []
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            cells.append(f"<td>{html_lib.escape(value)}</td>")
        if any(cell != "<td></td>" for cell in cells):
            rows.append(f"<tr><th>{row_index}</th>{''.join(cells)}</tr>")
    title = html_lib.escape(filename or "来源表格")
    sheet_name = html_lib.escape(sheet.title)
    return f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><title>{title}</title>
<style>body{{margin:0;color:#282b30;font:13px -apple-system,BlinkMacSystemFont,'PingFang SC',sans-serif}}
.sheet{{padding:16px;overflow:auto}}.caption{{margin-bottom:12px;color:#747970;font-size:12px}}
table{{border-collapse:collapse;background:#fff}}th,td{{min-width:90px;padding:7px 10px;border:1px solid #dde0d8;text-align:left;vertical-align:top;white-space:pre-wrap}}
th{{position:sticky;left:0;background:#f4f5f2;color:#6e8c19;font-weight:600}}thead th{{top:0}}</style></head>
<body><div class="sheet"><div class="caption">{title} · 工作表：{sheet_name}（最多显示 500 行、50 列）</div>
<table><tbody>{''.join(rows)}</tbody></table></div></body></html>"""


@app.get("/internal-api/health")
async def health() -> dict[str, Any]:
    return {
        "ok": True,
        "auth_mode": AUTH_MODE,
        "service_configured": bool(RAGFLOW_API_TOKEN),
        "assistant_count": len(ASSISTANT_CHAT_IDS),
    }


@app.get("/internal-api/auth/status")
async def auth_status(request: Request) -> dict[str, Any]:
    if AUTH_MODE == "trusted_lan":
        return {"authenticated": _in_trusted_network(request), "mode": AUTH_MODE}
    payload = _session_payload(request)
    return {
        "authenticated": bool(payload),
        "mode": AUTH_MODE,
        "user": payload.get("name") if payload else None,
    }


@app.get("/internal-api/conversations")
async def list_conversations(request: Request, assistant_id: str) -> JSONResponse:
    assistant_id = _history_assistant_id(assistant_id)
    owner_key, signed_cookie = _history_owner(request)
    response = JSONResponse({"items": history_store.list(owner_key, assistant_id)})
    return _set_history_cookie(response, signed_cookie)


@app.get("/internal-api/conversations/{conversation_id}")
async def get_conversation(request: Request, conversation_id: str) -> JSONResponse:
    owner_key, signed_cookie = _history_owner(request)
    record = history_store.get(owner_key, conversation_id)
    if record is None:
        raise HTTPException(status_code=404, detail="历史会话不存在")
    response = JSONResponse(record)
    return _set_history_cookie(response, signed_cookie)


@app.post("/internal-api/conversations")
async def save_conversation(request: Request) -> JSONResponse:
    body = await request.json()
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="历史会话格式不正确")
    assistant_id = _history_assistant_id(body.get("assistant_id"))
    messages = _safe_history_messages(body.get("messages"))
    session_id = body.get("session_id")
    if session_id is not None and (not isinstance(session_id, str) or len(session_id) > 255):
        raise HTTPException(status_code=400, detail="会话标识不正确")
    first_user_message = next((item["content"] for item in messages if item["role"] == "user"), "")
    title = body.get("title") if isinstance(body.get("title"), str) else first_user_message
    owner_key, signed_cookie = _history_owner(request)
    record = history_store.upsert(
        owner_key,
        assistant_id,
        body.get("id") if isinstance(body.get("id"), str) else None,
        session_id,
        title,
        messages,
    )
    response = JSONResponse(record)
    return _set_history_cookie(response, signed_cookie)


@app.get("/internal-api/operations/status")
async def operations_status(request: Request) -> dict[str, Any]:
    _require_access(request)
    return {
        "sync": sync_status(SYNC_STATUS_ROOT),
        "evaluation": evaluation_status(EVALUATION_STATUS_ROOT),
    }


@app.get("/internal-api/auth/login")
async def auth_login(request: Request, next: str = "/internal/") -> RedirectResponse:
    if AUTH_MODE != "wecom":
        return RedirectResponse("/internal/")
    if not WECOM_CORP_ID or not WECOM_AGENT_ID or not WECOM_REDIRECT_URI:
        raise HTTPException(status_code=503, detail="企业微信参数尚未配置")
    flow = _wecom_login_flow(request)
    state = serializer.dumps(
        {"nonce": secrets.token_urlsafe(16), "flow": flow, "next": _safe_next_path(next)}
    )
    if flow == "in_app":
        url = (
            "https://open.weixin.qq.com/connect/oauth2/authorize"
            f"?appid={quote(WECOM_CORP_ID)}&redirect_uri={quote(WECOM_REDIRECT_URI, safe='')}"
            f"&response_type=code&scope=snsapi_base&agentid={quote(WECOM_AGENT_ID)}"
            f"&state={quote(state)}#wechat_redirect"
        )
    else:
        url = (
            "https://login.work.weixin.qq.com/wwlogin/sso/login"
            f"?login_type=CorpApp&appid={quote(WECOM_CORP_ID)}&agentid={quote(WECOM_AGENT_ID)}"
            f"&redirect_uri={quote(WECOM_REDIRECT_URI, safe='')}&state={quote(state)}"
        )
    return RedirectResponse(url)


@app.get("/internal-api/auth/wecom/callback")
async def wecom_callback(request: Request, code: str = "", state: str = "") -> RedirectResponse:
    if AUTH_MODE != "wecom":
        return RedirectResponse("/internal/")
    if not code or not state:
        raise HTTPException(status_code=400, detail="企业微信回调参数不完整")
    try:
        state_payload = serializer.loads(state, max_age=600)
    except (BadSignature, SignatureExpired) as exc:
        raise HTTPException(status_code=400, detail="企业微信登录状态已失效") from exc
    if not isinstance(state_payload, dict):
        raise HTTPException(status_code=400, detail="企业微信登录状态无效")
    flow = state_payload.get("flow")
    if flow not in {"in_app", "qr"}:
        raise HTTPException(status_code=400, detail="企业微信登录方式无效")
    if not WECOM_AGENT_ID:
        raise HTTPException(status_code=503, detail="企业微信参数尚未配置")
    identity = await _wecom_user_identity(code, flow)
    identity["agent_id"] = WECOM_AGENT_ID
    response = RedirectResponse(_safe_next_path(str(state_payload.get("next", "/internal/"))))
    response.set_cookie(
        SESSION_COOKIE,
        serializer.dumps(identity),
        max_age=SESSION_MAX_AGE,
        httponly=True,
        samesite="lax",
        secure=COOKIE_SECURE,
        path="/",
    )
    return response


@app.post("/internal-api/auth/logout")
async def auth_logout() -> JSONResponse:
    response = JSONResponse({"ok": True})
    response.delete_cookie(SESSION_COOKIE, path="/")
    return response


@app.post("/internal-api/chat/completions")
async def chat_completions(request: Request) -> StreamingResponse:
    _require_access(request)
    body = await request.json()
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="请求格式不正确")
    assistant_id = body.get("assistant_id")
    if not isinstance(assistant_id, str):
        raise HTTPException(status_code=400, detail="缺少助手标识")
    payload = _safe_chat_request(body, assistant_id)

    async def stream() -> AsyncIterator[bytes]:
        timeout = httpx.Timeout(connect=15, read=None, write=30, pool=15)
        async with httpx.AsyncClient(timeout=timeout) as client:
            async with client.stream(
                "POST",
                f"{RAGFLOW_BASE_URL}/api/v1/chat/completions",
                headers={**_ragflow_headers(), "Content-Type": "application/json"},
                json=payload,
            ) as upstream:
                if upstream.status_code >= 400:
                    detail = (await upstream.aread()).decode("utf-8", errors="replace")
                    logger.warning("RAGFlow chat failed status=%s body=%s", upstream.status_code, detail[:500])
                    yield f"data: {json.dumps({'code': upstream.status_code, 'message': '知识助手暂时不可用'}, ensure_ascii=False)}\n\n".encode()
                    return
                async for chunk in upstream.aiter_bytes():
                    yield chunk

    return StreamingResponse(stream(), media_type="text/event-stream", headers={"Cache-Control": "no-cache"})


@app.get("/internal-api/documents/{doc_id}/preview")
async def document_preview(request: Request, doc_id: str, name: str = "", download: bool = False) -> Response:
    _require_access(request)
    timeout = httpx.Timeout(connect=15, read=120, write=30, pool=15)
    async with httpx.AsyncClient(timeout=timeout) as client:
        upstream = await client.get(
            f"{RAGFLOW_BASE_URL}/api/v1/documents/{quote(doc_id, safe='')}/preview",
            headers=_ragflow_headers(),
        )
    if upstream.status_code >= 400:
        logger.warning("RAGFlow preview failed status=%s doc_id=%s", upstream.status_code, doc_id)
        raise HTTPException(status_code=404, detail="来源文件不可用")

    extension = name.rsplit(".", 1)[-1].lower() if "." in name else ""
    if extension in {"xlsx", "xlsm"} and not download:
        try:
            return Response(content=_spreadsheet_html(upstream.content, name), media_type="text/html")
        except Exception as exc:
            logger.warning("Spreadsheet preview failed doc_id=%s: %s", doc_id, exc)

    response_headers = _document_response_headers(upstream.headers, name, download=download)
    return Response(content=upstream.content, status_code=200, headers=response_headers)
